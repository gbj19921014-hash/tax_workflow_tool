import json
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from workflow_service import build_country_payload, write_outputs


PROJECT_ROOT = Path(__file__).resolve().parent


class CountryWorkflowIntegrationTest(unittest.TestCase):
    def setUp(self):
        self.config = json.loads((PROJECT_ROOT / "config.json").read_text(encoding="utf-8"))

    def assert_detail_matches_csv(self, workbook, output_dir):
        csv_columns = list(pd.read_csv(output_dir / "命中明细.csv", nrows=0).columns)
        sheet = workbook["命中明细"]
        excel_columns = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
        self.assertEqual(excel_columns, csv_columns)

    def test_italy_april_regression_and_workbook(self):
        input_file = Path("/Users/dlab/Downloads/54961020579.csv")
        if not input_file.exists():
            self.skipTest("意大利回归 CSV 不存在")

        payload = build_country_payload(input_file, "2026-APR", "IT", self.config)
        self.assertEqual([row["命中行数"] for row in payload["summary"]], [3, 0, 0])
        self.assertEqual(payload["final"][0]["金额EUR"], 294.68)
        self.assertEqual(payload["final"][1]["金额EUR"], 53.14)

        with tempfile.TemporaryDirectory() as temp_dir:
            paths = write_outputs(payload, temp_dir)
            workbook = load_workbook(paths["xlsx_path"], data_only=False)
            self.assertEqual(workbook.sheetnames, ["最终输出", "命中明细"])
            self.assertEqual(workbook["最终输出"]["A1"].value, "意大利税务工作流输出")
            self.assert_detail_matches_csv(workbook, paths["output_dir"])

    def test_poland_february_regression_and_workbook(self):
        input_file = Path("/Users/dlab/Downloads/54849020552.csv")
        if not input_file.exists():
            self.skipTest("波兰回归 CSV 不存在")

        config = json.loads(json.dumps(self.config))
        config["exchange_rates_to_eur"]["2026-FEB"] = {
            "EUR": 1.0,
            "PLN": 0.23705895,
            "SEK": 0.0940284,
            "GBP": 1.14900927,
        }
        config["exchange_rate_source"] = "ECB月平均 2026-02"
        payload = build_country_payload(input_file, "2026-FEB", "PL", config)
        self.assertEqual([row["命中行数"] for row in payload["summary"]], [0, 0, 9, 10])
        self.assertEqual(payload["final"][0]["项目"], "3A命中EUR总和")
        self.assertEqual(payload["final"][0]["金额EUR"], 82.3)
        self.assertEqual(payload["final"][1]["公式"], "3A命中EUR总和/1.23*0.23")
        self.assertEqual(payload["final"][1]["金额EUR"], 15.39)

        with tempfile.TemporaryDirectory() as temp_dir:
            paths = write_outputs(payload, temp_dir)
            workbook = load_workbook(paths["xlsx_path"], data_only=False)
            self.assertEqual(workbook.sheetnames, ["最终输出", "命中明细"])
            self.assertEqual(workbook["最终输出"]["A1"].value, "波兰税务工作流输出")
            self.assert_detail_matches_csv(workbook, paths["output_dir"])

    def test_germany_q1_regression_and_workbook(self):
        input_file = Path("/Users/dlab/Desktop/新建文件夹/55149020613.csv")
        if not input_file.exists():
            self.skipTest("德国回归 CSV 不存在")

        config = json.loads(json.dumps(self.config))
        config["exchange_rates_to_eur"]["2026-Q1"] = {
            "EUR": 1.0,
            "PLN": None,
            "SEK": None,
            "GBP": None,
        }
        payload = build_country_payload(input_file, "2026-Q1", "DE", config)
        self.assertEqual([row["命中行数"] for row in payload["summary"]], [5, 0, 13, 230])
        self.assertEqual(
            [row["EUR金额总和"] for row in payload["summary"]],
            [70.57, 0.0, 1644.98, 20259.81],
        )
        self.assertEqual(payload["final"][0]["金额EUR"], 21975.36)
        self.assertEqual(payload["final"][1]["公式"], "A-D命中EUR总和/1.19*0.19")
        self.assertEqual(payload["final"][1]["金额EUR"], 3508.67)

        with tempfile.TemporaryDirectory() as temp_dir:
            paths = write_outputs(payload, temp_dir)
            workbook = load_workbook(paths["xlsx_path"], data_only=False)
            self.assertEqual(workbook.sheetnames, ["最终输出", "命中明细"])
            self.assertEqual(workbook["最终输出"]["A1"].value, "德国税务工作流输出")
            self.assertEqual(workbook["最终输出"]["A3"].value, "算税期间")
            self.assert_detail_matches_csv(workbook, paths["output_dir"])

    def test_germany_q3_manual_review_row_is_included_and_highlighted(self):
        input_file = Path("/Users/dlab/Downloads/7-9月德国税金订单，销售额10310.89.xlsx")
        if not input_file.exists():
            self.skipTest("德国Q3核对文件不存在")

        sheets = ["A部分B2C之EU-DE", "C部分B2B之DE-EU", "D部分本土B2B"]
        source = pd.concat(
            [pd.read_excel(input_file, sheet_name=sheet, dtype=str) for sheet in sheets],
            ignore_index=True,
        )
        config = json.loads(json.dumps(self.config))
        config["exchange_rates_to_eur"]["2025-Q3"] = {
            "EUR": 1.0,
            "PLN": None,
            "SEK": None,
            "GBP": None,
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            csv_path = Path(temp_dir) / "germany-q3.csv"
            source.to_csv(csv_path, index=False, encoding="utf-8-sig")
            payload = build_country_payload(csv_path, "2025-Q3", "DE", config)
            self.assertEqual([row["命中行数"] for row in payload["summary"]], [23, 0, 8, 163])
            self.assertEqual(payload["final"][0]["金额EUR"], 10310.89)
            self.assertEqual(payload["final"][1]["金额EUR"], 1646.28)

            review_rows = [row for row in payload["detail"] if row.get("人工确认提示")]
            self.assertEqual(len(review_rows), 1)
            self.assertEqual(review_rows[0]["BUYER_VAT_NUMBER"], "ESB40654055")
            self.assertEqual(float(review_rows[0]["TOTAL_ACTIVITY_VALUE_AMT_VAT_INCL"]), 249.99)

            paths = write_outputs(payload, temp_dir)
            workbook = load_workbook(paths["xlsx_path"], data_only=False)
            detail_sheet = workbook["命中明细"]
            headers = [detail_sheet.cell(1, column).value for column in range(1, detail_sheet.max_column + 1)]
            review_column = headers.index("人工确认提示") + 1
            highlighted = [
                row_index
                for row_index in range(2, detail_sheet.max_row + 1)
                if detail_sheet.cell(row_index, review_column).value
            ]
            self.assertEqual(len(highlighted), 1)
            self.assertEqual(detail_sheet.cell(highlighted[0], 1).fill.fgColor.rgb, "00FFF2CC")
            self.assert_detail_matches_csv(workbook, paths["output_dir"])

    def test_france_quarter_workflow_and_workbook(self):
        rows = [
            {
                "ACTIVITY_PERIOD": "2026-JAN",
                "TAX_COLLECTION_RESPONSIBILITY": "SELLER",
                "PRICE_OF_ITEMS_VAT_RATE_PERCENT": "20",
                "BUYER_VAT_NUMBER": "FR12345678901",
                "SALE_DEPART_COUNTRY": "FR",
                "SALE_ARRIVAL_COUNTRY": "DE",
                "TRANSACTION_CURRENCY_CODE": "EUR",
                "TOTAL_ACTIVITY_VALUE_AMT_VAT_INCL": "120",
            },
            {
                "ACTIVITY_PERIOD": "2026-FEB",
                "TAX_COLLECTION_RESPONSIBILITY": "SELLER",
                "PRICE_OF_ITEMS_VAT_RATE_PERCENT": "20",
                "BUYER_VAT_NUMBER": "",
                "SALE_DEPART_COUNTRY": "DE",
                "SALE_ARRIVAL_COUNTRY": "FR",
                "TRANSACTION_CURRENCY_CODE": "PLN",
                "TOTAL_ACTIVITY_VALUE_AMT_VAT_INCL": "246",
            },
            {
                "ACTIVITY_PERIOD": "2026-MAR",
                "TAX_COLLECTION_RESPONSIBILITY": "SELLER",
                "PRICE_OF_ITEMS_VAT_RATE_PERCENT": "",
                "BUYER_VAT_NUMBER": "",
                "SALE_DEPART_COUNTRY": "IT",
                "SALE_ARRIVAL_COUNTRY": "FR",
                "TRANSACTION_CURRENCY_CODE": "EUR",
                "TOTAL_ACTIVITY_VALUE_AMT_VAT_INCL": "36",
            },
            {
                "ACTIVITY_PERIOD": "2026-MAR",
                "TAX_COLLECTION_RESPONSIBILITY": "MARKETPLACE",
                "PRICE_OF_ITEMS_VAT_RATE_PERCENT": "20",
                "BUYER_VAT_NUMBER": "FR12345678901",
                "SALE_DEPART_COUNTRY": "FR",
                "SALE_ARRIVAL_COUNTRY": "DE",
                "TRANSACTION_CURRENCY_CODE": "EUR",
                "TOTAL_ACTIVITY_VALUE_AMT_VAT_INCL": "999",
            },
            {
                "ACTIVITY_PERIOD": "2026-MAR",
                "TAX_COLLECTION_RESPONSIBILITY": "SELLER",
                "PRICE_OF_ITEMS_VAT_RATE_PERCENT": "0",
                "BUYER_VAT_NUMBER": "",
                "SALE_DEPART_COUNTRY": "DE",
                "SALE_ARRIVAL_COUNTRY": "FR",
                "TRANSACTION_CURRENCY_CODE": "EUR",
                "TOTAL_ACTIVITY_VALUE_AMT_VAT_INCL": "888",
            },
        ]
        config = json.loads(json.dumps(self.config))
        config["exchange_rates_to_eur"]["2026-Q1"] = {
            "EUR": 1.0,
            "PLN": 0.5,
            "SEK": None,
            "GBP": None,
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            csv_path = Path(temp_dir) / "france-q1.csv"
            pd.DataFrame(rows).to_csv(csv_path, index=False, encoding="utf-8-sig")
            payload = build_country_payload(csv_path, "2026-Q1", "FR", config)
            self.assertEqual([row["命中行数"] for row in payload["summary"]], [3])
            self.assertEqual(payload["summary"][0]["EUR金额总和"], 279.0)
            self.assertEqual(payload["final"][0]["项目"], "自行缴税订单命中EUR总和")
            self.assertEqual(payload["final"][0]["金额EUR"], 279.0)
            self.assertEqual(payload["final"][1]["公式"], "自行缴税订单命中EUR总和/1.2*0.2")
            self.assertEqual(payload["final"][1]["金额EUR"], 46.5)

            paths = write_outputs(payload, temp_dir)
            workbook = load_workbook(paths["xlsx_path"], data_only=False)
            self.assertEqual(workbook.sheetnames, ["最终输出", "命中明细"])
            self.assertEqual(workbook["最终输出"]["A1"].value, "法国税务工作流输出")
            self.assert_detail_matches_csv(workbook, paths["output_dir"])

    def test_netherlands_quarter_workflow_and_workbook(self):
        rows = [
            {
                "ACTIVITY_PERIOD": "2026-JAN",
                "TAX_COLLECTION_RESPONSIBILITY": "MARKETPLACE",
                "PRICE_OF_ITEMS_VAT_RATE_PERCENT": "0",
                "BUYER_VAT_NUMBER": "",
                "SALE_DEPART_COUNTRY": "DE",
                "SALE_ARRIVAL_COUNTRY": "NL",
                "TRANSACTION_CURRENCY_CODE": "EUR",
                "TOTAL_ACTIVITY_VALUE_AMT_VAT_INCL": "121",
            },
            {
                "ACTIVITY_PERIOD": "2026-FEB",
                "TAX_COLLECTION_RESPONSIBILITY": "SELLER",
                "PRICE_OF_ITEMS_VAT_RATE_PERCENT": "0.21",
                "BUYER_VAT_NUMBER": "",
                "SALE_DEPART_COUNTRY": "GB",
                "SALE_ARRIVAL_COUNTRY": "NL",
                "TRANSACTION_CURRENCY_CODE": "PLN",
                "TOTAL_ACTIVITY_VALUE_AMT_VAT_INCL": "242",
            },
            {
                "ACTIVITY_PERIOD": "2026-MAR",
                "TAX_COLLECTION_RESPONSIBILITY": "SELLER",
                "PRICE_OF_ITEMS_VAT_RATE_PERCENT": "0.21",
                "BUYER_VAT_NUMBER": "",
                "SALE_DEPART_COUNTRY": "NL",
                "SALE_ARRIVAL_COUNTRY": "AT",
                "TRANSACTION_CURRENCY_CODE": "EUR",
                "TOTAL_ACTIVITY_VALUE_AMT_VAT_INCL": "60.5",
            },
            {
                "ACTIVITY_PERIOD": "2026-MAR",
                "TAX_COLLECTION_RESPONSIBILITY": "SELLER",
                "PRICE_OF_ITEMS_VAT_RATE_PERCENT": "0.21",
                "BUYER_VAT_NUMBER": "",
                "SALE_DEPART_COUNTRY": "NL",
                "SALE_ARRIVAL_COUNTRY": "DE",
                "TRANSACTION_CURRENCY_CODE": "EUR",
                "TOTAL_ACTIVITY_VALUE_AMT_VAT_INCL": "999",
            },
        ]
        config = json.loads(json.dumps(self.config))
        config["exchange_rates_to_eur"]["2026-Q1"] = {
            "EUR": 1.0,
            "PLN": 0.5,
            "SEK": None,
            "GBP": None,
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            csv_path = Path(temp_dir) / "netherlands-q1.csv"
            pd.DataFrame(rows).to_csv(csv_path, index=False, encoding="utf-8-sig")
            payload = build_country_payload(csv_path, "2026-Q1", "NL", config)
            self.assertEqual([row["项目"] for row in payload["summary"]], ["A", "B"])
            self.assertEqual([row["命中行数"] for row in payload["summary"]], [1, 1])
            self.assertEqual([row["EUR金额总和"] for row in payload["summary"]], [121.0, 60.5])
            self.assertEqual(payload["final"][0]["项目"], "Seller部分总销售额")
            self.assertEqual(payload["final"][0]["金额EUR"], 181.5)
            self.assertEqual(payload["final"][1]["公式"], "Seller部分总销售额/1.21*0.21")
            self.assertEqual(payload["final"][1]["金额EUR"], 31.5)

            paths = write_outputs(payload, temp_dir)
            workbook = load_workbook(paths["xlsx_path"], data_only=False)
            self.assertEqual(workbook.sheetnames, ["最终输出", "命中明细"])
            self.assertEqual(workbook["最终输出"]["A1"].value, "荷兰税务工作流输出")
            self.assert_detail_matches_csv(workbook, paths["output_dir"])

    def test_spain_quarter_workflow_and_workbook(self):
        rows = [
            {
                "ACTIVITY_PERIOD": "2026-JAN",
                "TRANSACTION_TYPE": "SALE",
                "TAX_COLLECTION_RESPONSIBILITY": "SELLER",
                "PRICE_OF_ITEMS_VAT_RATE_PERCENT": "0.22",
                "BUYER_VAT_NUMBER": "ESB12345678",
                "SALE_DEPART_COUNTRY": "ES",
                "SALE_ARRIVAL_COUNTRY": "FR",
                "TRANSACTION_CURRENCY_CODE": "EUR",
                "TOTAL_ACTIVITY_VALUE_AMT_VAT_INCL": "122",
            },
            {
                "ACTIVITY_PERIOD": "2026-FEB",
                "TRANSACTION_TYPE": "SALE",
                "TAX_COLLECTION_RESPONSIBILITY": "SELLER",
                "PRICE_OF_ITEMS_VAT_RATE_PERCENT": "0.22",
                "BUYER_VAT_NUMBER": "",
                "SALE_DEPART_COUNTRY": "FR",
                "SALE_ARRIVAL_COUNTRY": "ES",
                "TRANSACTION_CURRENCY_CODE": "PLN",
                "TOTAL_ACTIVITY_VALUE_AMT_VAT_INCL": "244",
            },
            {
                "ACTIVITY_PERIOD": "2026-MAR",
                "TRANSACTION_TYPE": "SALE",
                "TAX_COLLECTION_RESPONSIBILITY": "SELLER",
                "PRICE_OF_ITEMS_VAT_RATE_PERCENT": "",
                "BUYER_VAT_NUMBER": "",
                "SALE_DEPART_COUNTRY": "DE",
                "SALE_ARRIVAL_COUNTRY": "ES",
                "TRANSACTION_CURRENCY_CODE": "EUR",
                "TOTAL_ACTIVITY_VALUE_AMT_VAT_INCL": "61",
            },
            {
                "ACTIVITY_PERIOD": "2026-MAR",
                "TRANSACTION_TYPE": "COMMINGLING_BUY",
                "TAX_COLLECTION_RESPONSIBILITY": "SELLER",
                "PRICE_OF_ITEMS_VAT_RATE_PERCENT": "0.22",
                "BUYER_VAT_NUMBER": "ESB12345678",
                "SALE_DEPART_COUNTRY": "ES",
                "SALE_ARRIVAL_COUNTRY": "FR",
                "TRANSACTION_CURRENCY_CODE": "EUR",
                "TOTAL_ACTIVITY_VALUE_AMT_VAT_INCL": "999",
            },
        ]
        config = json.loads(json.dumps(self.config))
        config["exchange_rates_to_eur"]["2026-Q1"] = {
            "EUR": 1.0,
            "PLN": 0.5,
            "SEK": None,
            "GBP": None,
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            csv_path = Path(temp_dir) / "spain-q1.csv"
            pd.DataFrame(rows).to_csv(csv_path, index=False, encoding="utf-8-sig")
            payload = build_country_payload(csv_path, "2026-Q1", "ES", config)
            self.assertEqual([row["项目"] for row in payload["summary"]], ["1", "2", "3"])
            self.assertEqual([row["命中行数"] for row in payload["summary"]], [1, 1, 1])
            self.assertEqual([row["EUR金额总和"] for row in payload["summary"]], [122.0, 122.0, 61.0])
            self.assertEqual(payload["final"][0]["项目"], "1-3部分命中EUR总和")
            self.assertEqual(payload["final"][0]["金额EUR"], 305.0)
            self.assertEqual(payload["final"][1]["公式"], "1-3部分命中EUR总和/1.22*0.22")
            self.assertEqual(payload["final"][1]["金额EUR"], 55.0)

            paths = write_outputs(payload, temp_dir)
            workbook = load_workbook(paths["xlsx_path"], data_only=False)
            self.assertEqual(workbook.sheetnames, ["最终输出", "命中明细"])
            self.assertEqual(workbook["最终输出"]["A1"].value, "西班牙税务工作流输出")
            self.assert_detail_matches_csv(workbook, paths["output_dir"])

    def test_uk_shifted_quarter_workflow_and_workbook(self):
        rows = [
            {
                "ACTIVITY_PERIOD": "2026-FEB",
                "TAX_COLLECTION_RESPONSIBILITY": "MARKETPLACE",
                "BUYER_VAT_NUMBER": "",
                "SALE_DEPART_COUNTRY": "GB",
                "SALE_ARRIVAL_COUNTRY": "GB",
                "ARRIVAL_POST_CODE": "SW1A 1AA",
                "TRANSACTION_CURRENCY_CODE": "GBP",
                "TOTAL_ACTIVITY_VALUE_AMT_VAT_INCL": "120",
            },
            {
                "ACTIVITY_PERIOD": "2026-MAR",
                "TAX_COLLECTION_RESPONSIBILITY": "SELLER",
                "BUYER_VAT_NUMBER": "GB123456789",
                "SALE_DEPART_COUNTRY": "GB",
                "SALE_ARRIVAL_COUNTRY": "GB",
                "ARRIVAL_POST_CODE": "EC1A 1BB",
                "TRANSACTION_CURRENCY_CODE": "GBP",
                "TOTAL_ACTIVITY_VALUE_AMT_VAT_INCL": "240",
            },
            {
                "ACTIVITY_PERIOD": "2026-APR",
                "TAX_COLLECTION_RESPONSIBILITY": "SELLER",
                "BUYER_VAT_NUMBER": "",
                "SALE_DEPART_COUNTRY": "GB",
                "SALE_ARRIVAL_COUNTRY": "GB",
                "ARRIVAL_POST_CODE": "JE2 3AA",
                "TRANSACTION_CURRENCY_CODE": "GBP",
                "TOTAL_ACTIVITY_VALUE_AMT_VAT_INCL": "60",
            },
            {
                "ACTIVITY_PERIOD": "2026-APR",
                "TAX_COLLECTION_RESPONSIBILITY": "SELLER",
                "BUYER_VAT_NUMBER": "",
                "SALE_DEPART_COUNTRY": "GB",
                "SALE_ARRIVAL_COUNTRY": "GB",
                "ARRIVAL_POST_CODE": "GY8 0LP",
                "TRANSACTION_CURRENCY_CODE": "GBP",
                "TOTAL_ACTIVITY_VALUE_AMT_VAT_INCL": "36",
            },
            {
                "ACTIVITY_PERIOD": "2026-APR",
                "TAX_COLLECTION_RESPONSIBILITY": "SELLER",
                "BUYER_VAT_NUMBER": "",
                "SALE_DEPART_COUNTRY": "GB",
                "SALE_ARRIVAL_COUNTRY": "GB",
                "ARRIVAL_POST_CODE": "KY11 8ST",
                "TRANSACTION_CURRENCY_CODE": "GBP",
                "TOTAL_ACTIVITY_VALUE_AMT_VAT_INCL": "300",
            },
        ]
        config = json.loads(json.dumps(self.config))
        config["exchange_rates_to_eur"]["2026-Q1"] = {
            "EUR": 1.0,
            "PLN": None,
            "SEK": None,
            "GBP": 1.25,
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            csv_path = Path(temp_dir) / "uk-q1.csv"
            pd.DataFrame(rows).to_csv(csv_path, index=False, encoding="utf-8-sig")
            payload = build_country_payload(csv_path, "2026-Q1", "GB", config)
            self.assertEqual(payload["activity_periods"], ["2026-FEB", "2026-MAR", "2026-APR"])
            self.assertEqual([row["项目"] for row in payload["summary"]], ["b", "c1", "c2"])
            self.assertEqual([row["命中行数"] for row in payload["summary"]], [1, 2, 1])
            self.assertEqual([row["原币金额总和"] for row in payload["summary"]], [240.0, 96.0, 300.0])
            self.assertEqual([row["EUR金额总和"] for row in payload["summary"]], [300.0, 120.0, 375.0])
            self.assertEqual(payload["final"][0]["项目"], "零税率销售额")
            self.assertEqual(payload["final"][0]["金额GBP"], 96.0)
            self.assertEqual(payload["final"][1]["项目"], "应税命中GBP总和")
            self.assertEqual(payload["final"][1]["金额GBP"], 540.0)
            self.assertEqual(payload["final"][2]["公式"], "应税命中GBP总和/1.2*0.2")
            self.assertEqual(payload["final"][2]["金额GBP"], 90.0)

            paths = write_outputs(payload, temp_dir)
            workbook = load_workbook(paths["xlsx_path"], data_only=False)
            self.assertEqual(workbook.sheetnames, ["最终输出", "命中明细"])
            self.assertEqual(workbook["最终输出"]["A1"].value, "英国税务工作流输出")
            self.assert_detail_matches_csv(workbook, paths["output_dir"])


if __name__ == "__main__":
    unittest.main()

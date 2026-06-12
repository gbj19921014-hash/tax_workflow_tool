from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


TITLE_FILL = "17365D"
HEADER_FILL = "1F4E78"
LABEL_FILL = "D9EAF7"
REVIEW_FILL = "FFF2CC"


def _write_table(sheet, start_row, start_col, headers, rows):
    for offset, header in enumerate(headers):
        cell = sheet.cell(row=start_row, column=start_col + offset, value=header)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row_index, row in enumerate(rows, start=start_row + 1):
        for col_offset, header in enumerate(headers):
            value = row.get(header, "")
            cell = sheet.cell(row=row_index, column=start_col + col_offset, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def _autosize(sheet, max_width=60):
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        width = 10
        for cell in column_cells:
            if cell.value is None:
                continue
            text = str(cell.value)
            width = max(width, min(max_width, len(text) + 2))
        sheet.column_dimensions[column_letter].width = width


def _format_amount_columns(sheet, columns, start_row=2):
    for column in columns:
        for row in range(start_row, sheet.max_row + 1):
            sheet[f"{column}{row}"].number_format = "#,##0.00"


def build_workbook(data, output_path):
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    final_sheet = wb.active
    final_sheet.title = "最终输出"
    detail_sheet = wb.create_sheet("命中明细")

    final_sheet.sheet_view.showGridLines = False
    final_sheet.merge_cells("A1:F1")
    country_name = data.get("country_name", data.get("tax_country", ""))
    final_sheet["A1"] = f"{country_name}税务工作流输出"
    final_sheet["A1"].fill = PatternFill("solid", fgColor=TITLE_FILL)
    final_sheet["A1"].font = Font(bold=True, color="FFFFFF", size=16)

    info_rows = [
        ("算税期间", data.get("activity_period", "")),
        ("算税国家", data.get("tax_country", "")),
        ("源文件", data.get("source_file", "")),
        ("已注册税号国家", ", ".join(data.get("registered_vat_countries", []))),
        ("未注册欧盟国家", ", ".join(data.get("unregistered_eu_countries", []))),
        (
            "期间汇率",
            ", ".join(
                f"{key}:{value if value is not None else '未填'}"
                for key, value in data.get("rates_for_period", data.get("rates_for_month", {})).items()
            ),
        ),
        ("汇率来源", data.get("exchange_rate_source", "")),
    ]
    for row_index, (label, value) in enumerate(info_rows, start=3):
        final_sheet.cell(row=row_index, column=1, value=label).fill = PatternFill("solid", fgColor=LABEL_FILL)
        final_sheet.cell(row=row_index, column=1).font = Font(bold=True)
        final_sheet.cell(row=row_index, column=2, value=value).alignment = Alignment(wrap_text=True)

    summary_start = 10
    summary_headers = ["项目", "名称", "命中行数", "EUR金额总和", "税金EUR", "缺少汇率行数"]
    _write_table(final_sheet, summary_start, 1, summary_headers, data.get("summary", []))
    final_start = summary_start + len(data.get("summary", [])) + 3
    _write_table(final_sheet, final_start, 1, ["项目", "公式", "金额EUR", "缺少汇率行数"], data.get("final", []))
    _format_amount_columns(final_sheet, ["D", "E"], summary_start + 1)
    _format_amount_columns(final_sheet, ["C"], final_start + 1)
    final_sheet.freeze_panes = "A11"

    detail_rows = data.get("detail", [])
    detail_headers = []
    for row in detail_rows:
        for header in row:
            if header not in detail_headers:
                detail_headers.append(header)
    _write_table(detail_sheet, 1, 1, detail_headers, detail_rows)
    if "人工确认提示" in detail_headers:
        review_column = detail_headers.index("人工确认提示") + 1
        review_fill = PatternFill("solid", fgColor=REVIEW_FILL)
        for row_index in range(2, detail_sheet.max_row + 1):
            if detail_sheet.cell(row_index, review_column).value:
                for column_index in range(1, detail_sheet.max_column + 1):
                    detail_sheet.cell(row_index, column_index).fill = review_fill
    detail_sheet.freeze_panes = "A2"

    for sheet in wb.worksheets:
        _autosize(sheet)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(output_path)
    return output_path
